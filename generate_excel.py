import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

excel_path = r'D:\Downloads\rekap_pendaftaran_pikr.xlsx'
df_orig = pd.read_excel(excel_path, skiprows=2)
df_orig.columns = [str(c).strip() for c in df_orig.columns]
df = df_orig.dropna(subset=['Nama']).copy()

# Detailed mapping
# (PNG Row #, PNG Name, PNG Class, Excel Row Index 0-based, Typo Note)
mapping = [
    (1, 'Nadia Al-Hafizho', 'X.2', 23, 'Nama di presensi "Nadia Al-Hafizho", nama resmi pendaftar "Nadira Al Hafizho Selimat"'),
    (2, 'Syakira Atiqah Ikhansa Novacia', 'XI.2', 66, 'Nama di presensi "Syakira Atiqah Ikhansa Novacia", nama resmi pendaftar "Syakira Ariqah Khansa Novalia"'),
    (3, 'Zahra Wahyuni', 'X.6', 49, 'Nama di presensi "Zahra Wahyuni", nama resmi pendaftar "Zahira Wahyuni"'),
    (4, 'Naila Oratia Dwiani', 'X.8', 62, 'Nama di presensi "Naila Oratia Dwiani" (X.8), nama resmi pendaftar "Naila Dratia Dwiani" (X-9)'),
    (5, 'Aletya Dwi Putri', 'X.6', 46, 'Nama di presensi "Aletya Dwi Putri", nama resmi pendaftar "Altya Dwi Putri"'),
    (6, 'Mishelni Hafizah', 'X.7', 54, 'Nama sesuai'),
    (7, 'Elisya Dwi Fitriani', 'X.2', 16, 'Nama sesuai'),
    (8, 'Cecilia Audia Merzyta', 'X.2', 15, 'Nama sesuai'),
    (9, 'Nabila Ayu Lestari', 'X.2', 21, 'Nama di presensi "Nabila Ayu Lestari", nama resmi pendaftar "Nabila Ayu Lastari"'),
    (10, 'Rani Sebrina', 'X.9', 63, 'Nama sesuai'),
    (11, 'Kaya Salsabila', 'X.9', 61, 'Nama di presensi "Kaya Salsabila", nama resmi pendaftar "Kayla Salsabila"'),
    (12, 'Nadiya Al-Hafizhoh', 'X.1', 6, 'Nama sesuai'),
    (13, 'Nayyara Nur Artanty', 'X.4', 33, 'Nama sesuai'),
    (14, 'Naila Izzah Sabilillah', 'X.4', 31, 'Nama sesuai'),
    (15, 'Adrinne Anandya Arsya', 'X.2', 12, 'Nama sesuai'),
    (16, 'Yasmin Qanitah', 'X.2', 8, 'Kelas di presensi X.2, kelas resmi di Excel X-1'),
    (17, 'Zalwa Nur Afifa Putri', 'X.1', 10, 'Nama sesuai'),
    (18, 'Ghisa Fawzah Putri', 'X.1', 2, 'Nama sesuai'),
    (19, 'Naira Khoirunnisa', 'X.1', 7, 'Nama sesuai'),
    (20, 'Ahza Shoagurdeia', 'X.1', 11, 'Nama di presensi "Ahza Shoagurdeia", nama resmi pendaftar "Ahza Sholaqurdelia"'),
    (21, 'Inayah Salsabilah Tanjung', 'X.2', 18, 'Nama di presensi "Inayah Salsabilah Tanjung", nama resmi pendaftar "Inayah Salsabillah Tanjung"'),
    (22, 'Annisa Almadinah', 'X.2', 13, 'Nama sesuai'),
    (23, 'Chika Azzah Afifah', 'X.1', 0, 'Nama sesuai'),
    (24, 'Nurul Dhiya Niffah', 'XI.1', 65, 'Nama di presensi "Nurul Dhiya Niffah", nama resmi pendaftar "Aurel Dhiya Rifqah"'),
    (25, 'Naya Fitri Ramadhani', 'X.4', 32, 'Nama di presensi "Naya Fitri Ramadhani", nama resmi pendaftar "Nayla Fitri Ramadhani"'),
    (26, 'Windi Catur Ratri', 'X.4', 36, 'Nama sesuai'),
    (27, 'Wsnda Naisa Revita', 'X.5', 41, 'Nama di presensi "Wsnda Naisa Revita", nama resmi pendaftar "Wanda Naira Revita"'),
    (28, 'Anggraini Jenia Novianti', 'X.3', 27, 'Nama di presensi "Anggraini Jenia Novianti", nama resmi pendaftar "Anggraini Junia Novianti"'),
    (29, 'Putri Khoirpa Chotunnisa', 'X.3', 30, 'Nama di presensi "Putri Khoirpa Chotunnisa", nama resmi pendaftar "Putri Kholipa Choironisa"'),
    (30, 'Okta Regina Putri', 'X.9', 34, 'Nama di presensi "Okta Regina Putri" (X.9), nama resmi pendaftar "Oktia Regina Putri" (X-4)'),
    (31, 'Siipa Enselina', 'X.2', 26, 'Nama di presensi "Siipa Enselina", nama resmi pendaftar "Silpa Enjelita"'),
    (32, 'Aisyah Munawaroh', 'X.6', 44, 'Nama sesuai'),
    (33, 'Adila Bilqis', 'X.6', 43, 'Nama di presensi "Adila Bilqis", nama resmi pendaftar "Adilah Bilqis"'),
    (34, 'Deby Adella Herlino', 'X.6', 47, 'Nama di presensi "Deby Adella Herlino", nama resmi pendaftar "Deby Adelia Herlino"'),
    (35, 'Ririn Utra Lestari', 'X.7', 56, 'Nama di presensi "Ririn Utra Lestari", nama resmi pendaftar "Ririn Citra Lestari"'),
    (36, 'Alpina Puspitasari', 'X.6', 45, 'Nama sesuai'),
    (37, 'Evi Rahmadhani', 'X.1', 1, 'Nama sesuai'),
    (38, 'Aisyah Denaka', 'X.7', 50, 'Nama di presensi "Aisyah Denaka", nama resmi pendaftar "Aisyah Denata"'),
    (39, 'Kareni Laura Rahmadani', 'X.7', 52, 'Nama sesuai'),
    (40, 'Mellani Ananta', 'X.7', 53, 'Nama sesuai'),
    (41, 'Janeeta', 'X.3', 29, 'Nama sesuai'),
    (42, 'Kanista Syarifatulzzahra', 'X.8', 58, 'Nama di presensi "Kanista Syarifatulzzahra", nama resmi pendaftar "Kalista Syarifatuzzahra"'),
    (43, 'Muhammad Amirul Amin', 'X.1', 5, 'Nama sesuai'),
    (44, 'M. Berlian Almirafi', 'X.1', 3, 'Nama sesuai'),
    (45, 'Yusuf Athallah', 'X.7', 9, 'Kelas di presensi X.7, kelas resmi di Excel X-1'),
    (46, 'Kenzie Favian Mahardika Ramaay', 'X.8', 59, 'Nama di presensi "Kenzie Favian Mahardika Ramaay", nama resmi pendaftar "Kenzie Favian Mahardika Ramaqy"'),
    (47, 'Kioshi Azka Azzukra', 'X.9', 64, 'Nama di presensi "Kioshi Azka Azzukra", nama resmi pendaftar "Kioshi Azka Azzikra"'),
    (48, 'Rifqi Al Habsyi', 'X.5', 40, 'Nama sesuai'),
    (49, 'M. Aziz Al Hady', 'X.2', 19, 'Nama sesuai'),
    (50, 'Muhammad Iqbal Syahputra', 'X.1', 4, 'Nama di presensi "Muhammad Iqbal Syahputra", nama resmi pendaftar "Muhammad Arzal Sya\'bani"'),
    (51, 'Ahmad Bagas Pratama', 'XI.5', 70, 'Nama di presensi "Ahmad Bagas Pratama", nama resmi pendaftar "Rahmad Bagas Pratama"')
]

hadir_dict = {item[3]: item for item in mapping}

# Build dataset
rekap_rows = []
for idx, row in df.iterrows():
    no = row['No.']
    nisn = str(row['NISN'])
    nama_resmi = row['Nama']
    kelas = row['Kelas']
    jk = row['Jenis Kelamin']
    wa = row['No. WhatsApp']
    email = row['Email']
    
    if idx in hadir_dict:
        status = 'Hadir'
        png_item = hadir_dict[idx]
        nama_png = png_item[1]
        kelas_png = png_item[2]
        keterangan = png_item[4]
    else:
        status = 'Tidak Hadir'
        nama_png = '-'
        kelas_png = '-'
        keterangan = '-'
        
    rekap_rows.append({
        'No': no,
        'NISN': nisn,
        'Nama Resmi (Excel)': nama_resmi,
        'Nama Presensi (PNG)': nama_png,
        'Kelas Pendaftaran': kelas,
        'Kelas Presensi': kelas_png,
        'Jenis Kelamin': jk,
        'No. WhatsApp': wa,
        'Status Kehadiran': status,
        'Keterangan Koreksi / Typo': keterangan
    })

df_rekap = pd.DataFrame(rekap_rows)

# Create Workbook with openpyxl
wb = openpyxl.Workbook()
# remove default sheet
wb.remove(wb.active)

# Styles
font_title = Font(name='Calibri', size=16, bold=True, color='1F4E79')
font_subtitle = Font(name='Calibri', size=11, italic=True, color='595959')
font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
font_data = Font(name='Calibri', size=11)
font_bold = Font(name='Calibri', size=11, bold=True)

fill_header = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
fill_hadir = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Soft Green
fill_tidak_hadir = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Soft Red/Orange
fill_typo = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Soft Yellow

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')

def build_sheet(ws, title, subtitle, headers, data_rows, is_rekap=False, is_typo=False):
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.append([title])
    ws.cell(row=1, column=1).font = font_title
    ws.append([subtitle])
    ws.cell(row=2, column=1).font = font_subtitle
    ws.append([]) # Blank row
    
    # Header Row
    ws.append(headers)
    header_row_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    
    # Data Rows
    for r_idx, row_data in enumerate(data_rows, start=5):
        ws.append(row_data)
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_data
            cell.border = thin_border
            
            # Alignments
            val_str = str(cell.value)
            if c_idx in [1, 2, 5, 6, 7, 9]: # No, NISN, Kelas, JK, Status
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Formatting highlight
            if 'Status Kehadiran' in headers:
                status_col_idx = headers.index('Status Kehadiran') + 1
                status_val = ws.cell(row=r_idx, column=status_col_idx).value
                if status_val == 'Hadir':
                    if c_idx == status_col_idx:
                        cell.fill = fill_hadir
                        cell.font = Font(name='Calibri', size=11, bold=True, color='375623')
                elif status_val == 'Tidak Hadir':
                    if c_idx == status_col_idx:
                        cell.fill = fill_tidak_hadir
                        cell.font = Font(name='Calibri', size=11, bold=True, color='C65911')
            
            if is_typo:
                cell.fill = fill_typo
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:
                continue
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Sheet 1: Rekap Keseluruhan
ws1 = wb.create_sheet(title='Rekap Keseluruhan')
headers1 = list(df_rekap.columns)
data1 = df_rekap.values.tolist()
build_sheet(ws1, 'REKAP KEHADIRAN PENDAFTAR PIK-R MANSEKU', 'Data Gabungan Hasil Matching Presensi PNG dengan Rekap Excel Pendaftaran', headers1, data1, is_rekap=True)

# Sheet 2: Tidak Hadir
ws2 = wb.create_sheet(title='Tidak Hadir (20 Siswa)')
df_tidak_hadir = df_rekap[df_rekap['Status Kehadiran'] == 'Tidak Hadir'].copy()
df_tidak_hadir['No'] = range(1, len(df_tidak_hadir) + 1)
df_tidak_hadir_clean = df_tidak_hadir[['No', 'NISN', 'Nama Resmi (Excel)', 'Kelas Pendaftaran', 'Jenis Kelamin', 'No. WhatsApp', 'Status Kehadiran']]
headers2 = list(df_tidak_hadir_clean.columns)
data2 = df_tidak_hadir_clean.values.tolist()
build_sheet(ws2, 'DAFTAR PENDAFTAR PIK-R YANG TIDAK HADIR', 'Total: 20 Siswa Terdaftar Tidak Hadir Pada Kegiatan Hari Ini', headers2, data2)

# Sheet 3: Hadir
ws3 = wb.create_sheet(title='Hadir (51 Siswa)')
df_hadir = df_rekap[df_rekap['Status Kehadiran'] == 'Hadir'].copy()
df_hadir['No'] = range(1, len(df_hadir) + 1)
df_hadir_clean = df_hadir[['No', 'NISN', 'Nama Resmi (Excel)', 'Nama Presensi (PNG)', 'Kelas Pendaftaran', 'Kelas Presensi', 'Jenis Kelamin', 'Status Kehadiran', 'Keterangan Koreksi / Typo']]
headers3 = list(df_hadir_clean.columns)
data3 = df_hadir_clean.values.tolist()
build_sheet(ws3, 'DAFTAR SISWA HADIR HARI INI', 'Total: 51 Siswa Hadir Berdasarkan Catatan Presensi (PNG)', headers3, data3)

# Sheet 4: Koreksi Typo
ws4 = wb.create_sheet(title='Koreksi Typo Nama (27 Siswa)')
df_typo = df_rekap[df_rekap['Keterangan Koreksi / Typo'].str.contains('presensi')].copy()
df_typo['No'] = range(1, len(df_typo) + 1)
df_typo_clean = df_typo[['No', 'NISN', 'Nama Resmi (Excel)', 'Nama Presensi (PNG)', 'Kelas Pendaftaran', 'Kelas Presensi', 'Keterangan Koreksi / Typo']]
headers4 = list(df_typo_clean.columns)
data4 = df_typo_clean.values.tolist()
build_sheet(ws4, 'DAFTAR KOREKSI PENULISAN NAMA / TYPO', 'Detail Penyesuaian Nama Antara Daftar Presensi (PNG) dan Data Resmi Pendaftaran (Excel)', headers4, data4, is_typo=True)

# Save output to D:\Downloads\ and e:\PIK-R-MANSEKU\
out_path_downloads = r'D:\Downloads\rekap_kehadiran_pikr_terupdate.xlsx'
out_path_workspace = r'e:\PIK-R-MANSEKU\rekap_kehadiran_pikr_terupdate.xlsx'

wb.save(out_path_downloads)
wb.save(out_path_workspace)

print(f"Successfully saved updated Excel to:\n1. {out_path_downloads}\n2. {out_path_workspace}")
